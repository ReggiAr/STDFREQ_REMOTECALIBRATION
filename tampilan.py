'''
SOFTWARE TIME FREQUENCY REMOTE CALIBRATION
CGGTTS ANALYZER
GUI--
UPDATE 02/08/2024
'''

# -------- Libraries

# -------- -------- System Libraries
import sys
import os

# -------- -------- GUI Libraries
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QProgressBar, QTextEdit, QDialog
from PyQt5.QtWidgets import QFileDialog, QHBoxLayout, QLineEdit, QPushButton, QComboBox, QSpinBox, QMessageBox, QSpacerItem, QSizePolicy
from PyQt5.QtGui import QPixmap, QFont, QIcon,  QDesktopServices
from PyQt5.QtCore import Qt, QUrl, pyqtSignal

# -------- -------- Time Libraries
from datetime import date
import datetime as dt
from astropy.time import Time

# -------- -------- Time Libraries
from openpyxl import Workbook
import numpy as np

# -------- GUI Design

# -------- -------- Color
oren = "background-color: #369FFF; color: white"

# -------- -------- Font
font = QFont("Inter",10)

# Class
class jendelautama(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()
    
    # ini berisi tampilan untuk bagian header
    def header(self):

        judul = QLabel(self)
        judul.setText("  Time and Frequency Remote Calibration Software\t\t\t\t\n  CGGTTS Analyzer")
        judul.setStyleSheet(f"{oren}; font-weight: DemiBold")
        judul.setFont(QFont("Inter", 12))

        spasi = QSpacerItem(60, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)

        self.informasi = QPushButton(self)
        self.informasi.setText(" i ")
        self.informasi.setStyleSheet(oren)
        self.informasi.setFont(font)

        # set tanggal hari
        hari_ini = date.today()
        tanggal_iso = hari_ini.isoformat()
        time_obj = Time(tanggal_iso, format='iso')
        tanggal_terformat = hari_ini.strftime("%d-%m-%Y")
        mjd = time_obj.mjd
        mjd_str = str(mjd)
        mjd_str = round(mjd)
        mjd_int = int(mjd_str)

        tanggalBiasa = QLabel(self)
        tanggalBiasa.setText(f"   {tanggal_terformat}   ")
        tanggalBiasa.setFont(font)
        tanggalBiasa.setStyleSheet(oren)

        tanggalMJD = QLabel(self)
        tanggalMJD.setText(f"    {mjd_int}    ")
        tanggalMJD.setFont(font)
        tanggalMJD.setStyleSheet(oren)

        satu = QHBoxLayout ()
        satu.addWidget(tanggalBiasa)
        satu.addWidget(tanggalMJD)
        satu.addItem(spasi)
        satu.addItem(spasi)
        satu.addWidget(self.informasi)

        headerLayout = QVBoxLayout ()
        headerLayout.addWidget(judul)
        headerLayout.addLayout(satu)
        headerLayout.addItem(spasi)

        return headerLayout
    
    # Untuk input input
    def input (self):

        nama = QLabel("Client Name")
        nama.setFont(font)
        tanggal = QLabel("MJD")
        tanggal.setFont(font)

        self.dirStandar = QPushButton (self)
        self.dirStandar.setText("Directory Standard")
        self.dirStandar.setFont(font)
        self.dirStandar.setStyleSheet(oren)

        self.dirUUT = QPushButton (self)
        self.dirUUT.setText("Directory UUT")
        self.dirUUT.setFont(font)
        self.dirUUT.setStyleSheet(oren)

        self.dirOutput = QPushButton (self)
        self.dirOutput.setText("Directory Output")
        self.dirOutput.setFont(font)
        self.dirOutput.setStyleSheet(oren)

        self.locStandar = QLineEdit(self)
        self.locStandar.setFont(font)
        self.locStandar.setReadOnly(True)
        self.locUUT = QLineEdit(self)
        self.locUUT.setFont(font)
        self.locUUT.setReadOnly(True)
        self.locOutput = QLineEdit(self)
        self.locOutput.setFont(font)
        self.locOutput.setReadOnly(True)

        self.tipeStandar = QComboBox(self)
        self.tipeStandar.setFont(font)
        self.tipeStandar.setStyleSheet(oren)
        self.tipeStandar.addItems(["01","2E"])

        self.tipeUUT = QComboBox(self)
        self.tipeUUT.setFont(font)
        self.tipeUUT.setStyleSheet(oren)
        self.tipeUUT.addItems(["01","2E"])

        spasi = QSpacerItem(60, 40, QSizePolicy.Expanding, QSizePolicy.Minimum)
        
        satu = QVBoxLayout()
        satu.addWidget(self.dirStandar)
        satu.addWidget(self.dirUUT)
        satu.addWidget(self.dirOutput)

        dua = QVBoxLayout()
        dua.addWidget(self.locStandar)
        dua.addWidget(self.locUUT)
        dua.addWidget(self.locOutput)

        tiga = QVBoxLayout()
        tiga.addWidget(self.tipeStandar)
        tiga.addWidget(self.tipeUUT)
        tiga.addItem(spasi)

        inputsLayout = QHBoxLayout()
        inputsLayout.addLayout(satu)
        inputsLayout.addLayout(dua)
        inputsLayout.addLayout(tiga)

        inputsLayout.setStretchFactor(satu,1)
        inputsLayout.setStretchFactor(dua,5)
        inputsLayout.setStretchFactor(tiga,1)

        return inputsLayout

    # Untuk running
    def run (self):
        satu = QHBoxLayout()
        dua = QHBoxLayout()
        tiga = QHBoxLayout()
        empat = QHBoxLayout()

        cl = QLabel()
        cl.setText("Client Name")
        cl.setFont(font)
        mjdl = QLabel()
        mjdl = QLabel()
        mjdl.setText("MJD")
        mjdl.setFont(font)
        utclabel = QLabel()
        utclabel.setText("UTC(IDN)")
        utclabel.setAlignment(Qt.AlignCenter)
        utclabel.setFont(font)
        uutlabel = QLabel()
        uutlabel.setAlignment(Qt.AlignCenter)
        uutlabel.setText("UUT")
        uutlabel.setFont(font)

        self.clientname = QLineEdit(self)
        self.mjdname = QLineEdit(self)

        self.utcname = QComboBox(self)
        self.utcname.setStyleSheet(oren)
        self.utcname.setFont(font)
        self.uutname = QComboBox(self)
        self.uutname.setStyleSheet(oren)
        self.uutname.setFont(font)

        spasi = QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.analize = QPushButton(self)
        self.analize.setText("ANALYZE")
        self.analize.setFont(font)
        self.analize.setStyleSheet(oren)

        satu.addWidget(cl)
        satu.addWidget(self.clientname)
        satu.addWidget(mjdl)
        satu.addWidget(self.mjdname)

        dua.addWidget(utclabel)
        dua.addWidget(uutlabel)

        tiga.addWidget(self.utcname)
        tiga.addWidget(self.uutname)

        empat.addItem(spasi)
        empat.addWidget(self.analize)

        run_layout = QVBoxLayout()
        run_layout.addItem(spasi)
        run_layout.addLayout(satu)
        run_layout.addLayout(dua)
        run_layout.addLayout(tiga)
        run_layout.addLayout(empat)

        return run_layout

    # untuk output
    def output (self):

        alllabel = QLabel()
        alllabel.setText("Allan Variance")
        alllabel.setFont(font)

        corrlabel = QLabel()
        corrlabel.setText("Correction")
        corrlabel.setFont(font)

        outlabel = QLabel()
        outlabel.setText("O U T P U T")
        outlabel.setFont(font)

        self.allan = QLineEdit(self)
        self.correction = QLineEdit(self)
        self.allan.setFont(font)
        self.correction.setFont(font)
        self.allan.setReadOnly(True)
        self.correction.setReadOnly(True)

        self.outputs = QTextEdit(self)
        self.outputs.setFont(font)

        self.delete = QPushButton(self)
        self.delete.setText("Delete")
        self.delete.setFont(font)
        self.delete.setStyleSheet(oren)

        self.uPsudorange = QPushButton(self)
        self.uPsudorange.setText("uPsudorange")
        self.uPsudorange.setFont(font)
        self.uPsudorange.setStyleSheet(oren)
        self.uPsudorange.clicked.connect(self.uPseudorange)

        spasi = QSpacerItem(60, 100, QSizePolicy.Expanding, QSizePolicy.Minimum)

        satu = QVBoxLayout()
        dua = QVBoxLayout()
        untukoutput = QHBoxLayout()

        satu.addWidget(alllabel)
        satu.addWidget(self.allan)
        satu.addWidget(corrlabel)
        satu.addWidget(self.correction)
        satu.addWidget(self.delete)
        satu.addWidget(self.uPsudorange)
        satu.addItem(spasi)

        dua.addWidget(outlabel)
        dua.addWidget(self.outputs)

        untukoutput.addLayout(satu)
        untukoutput.addLayout(dua)

        untukoutput.setStretchFactor(satu,1)
        untukoutput.setStretchFactor(dua,5)

        return untukoutput

    # footer
    def footer (self):
        foot = QVBoxLayout()

        spasi = QSpacerItem(30, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        kaki = QLabel()
        kaki.setText("CGGTTS Analyzer for Cesium Atomic Clock Remote Calibration © SNSU Time and Frequency Laboratory 2024 ")

        foot.addItem(spasi)
        foot.addWidget(kaki)

        return foot

    #---------------------- uPseudorange
    #GUI
    def uPseudorange (self):
        uPseudoBox = QDialog(self)
        uPseudoBox.setWindowTitle("uPseudorange Calculator")

        judul = QLabel(uPseudoBox)
        judul.setText(" uPseudorange Calculator\t")
        judul.setFont(font)
        judul.setStyleSheet(f"{oren}; font-weight: Bold")

        tutor = QLabel(uPseudoBox)
        tutor.setText("Prepare a folder containing CGGTTS data for a full month in .txt format, with files named from 1 to 30/31")
        tutor.setFont(font)
        tutor.setStyleSheet(oren)

        cfolder = QPushButton(uPseudoBox)
        cfolder.setText("Choose Folder")
        cfolder.setFont(font)
        cfolder.setStyleSheet(oren)
        cfolder.clicked.connect(self.showDialog)

        self.foldir = QLineEdit(uPseudoBox)
        self.foldir.setFont(font)
        self.foldir.setReadOnly(True)

        jumlahlabel = QLabel(uPseudoBox)
        jumlahlabel.setText("Total Days")
        jumlahlabel.setFont(font)

        self.jumlah = QComboBox(uPseudoBox)
        self.jumlah.addItems(["28","29","30","31"])
        self.jumlah.setFont(font)
        self.jumlah.setStyleSheet(oren)

        cal = QPushButton(uPseudoBox)
        cal.setText("Calculate")
        cal.setFont(font)
        cal.setStyleSheet(oren)
        cal.clicked.connect(self.calUPseudo)

        self.PseudoOutput = QLineEdit(uPseudoBox)
        self.PseudoOutput.setFont(font)
        self.PseudoOutput.setReadOnly(True)

        snumber = QLabel(uPseudoBox)
        snumber.setText("Sampling Number")
        snumber.setFont(font)

        self.samplingnumber = QLineEdit(uPseudoBox)
        self.samplingnumber.setFont(font)

        o = QLabel(uPseudoBox)
        o.setText("|")
        o.setFont(font)

        slabel = QLabel(uPseudoBox)
        slabel.setText("STDEV")
        slabel.setFont(font)

        ulabel = QLabel(uPseudoBox)
        ulabel.setText("uPseudorange")
        ulabel.setFont(font)

        self.uPs = QLineEdit(uPseudoBox)
        self.uPs.setFont(font)
        self.uPs.setReadOnly(True)

        spasi = QSpacerItem(30, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        kaki = QLabel(uPseudoBox)
        kaki.setText("CGGTTS Analyzer for Cesium Atomic Clock Remote Calibration © SNSU Time and Frequency Laboratory 2024 ")
        kaki.setFont(font)

        satu = QHBoxLayout()
        satu.addWidget(cfolder)
        satu.addWidget(self.foldir)
        satu.addWidget(jumlahlabel)
        satu.addWidget(self.jumlah)
        satu.setStretchFactor(cfolder,1)
        satu.setStretchFactor(self.foldir,5)
        satu.setStretchFactor(jumlahlabel,1)
        satu.setStretchFactor(self.jumlah,1)

        dua = QHBoxLayout()
        dua.addWidget(cal)
        dua.addWidget(snumber)
        dua.addWidget(self.samplingnumber)
        dua.addWidget(o)
        dua.addWidget(slabel)
        dua.addWidget(self.PseudoOutput)
        dua.addWidget(ulabel)
        dua.addWidget(self.uPs)

        layout = QVBoxLayout(uPseudoBox)
        layout.addWidget(judul)
        layout.addItem(spasi)
        layout.addWidget(tutor)
        layout.addItem(spasi)
        layout.addLayout(satu)
        layout.addLayout(dua)
        layout.addItem(spasi)
        layout.addWidget(kaki)

        uPseudoBox.exec_()

    #----------Buka File
    def showDialog(self):
        # Buka dialog pemilihan folder
        folder = QFileDialog.getExistingDirectory(self, 'Pilih Folder')

        # Jika pengguna memilih folder, tampilkan alamat folder di QLineEdit
        if folder:
            self.foldir.setText(folder)

    # ----------- Perhitungan
    def calUPseudo(self):
        
        days = self.jumlah.currentText()
        day = int(days)
        workbook = Workbook()
        worksheet = workbook.create_sheet(title=f" Data CGGTTS ")

        stdev = []
        try:
            for a in range (day):
                refgps = []
                self.PseudoOutput.setText(f"Calculate {a+1} day")
                #Read File
                with open(f'{self.foldir.text()}/{a+1}.txt', 'r') as Filestd:
                    # Baca semua baris kecuali baris pertama
                    cggtts = Filestd.readlines()

                nama_kolom = ['SAT','STTIME','REFSYS']
                header = cggtts[17].strip().split()
                std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]

                # Loop melalui setiap baris file
                for line in cggtts[19:]:  # Mulai dari baris kedelapan belas karena baris pertama adalah header
                    data = line.strip().split()
        
                    # Pastikan jumlah elemen dalam baris sesuai dengan jumlah kolom yang diharapkan
                    if len(data) >= max(std_indeks_kolom) + 1:

                        # Masukkan ke dalam list yang sesuai
                        if data[std_indeks_kolom[2]].isdigit() or data[std_indeks_kolom[2]].startswith("-") or data[std_indeks_kolom[2]].startswith("+"):
                            refgps.append(data[std_indeks_kolom[2]])
                        else:
                            print("Skipping stdRefGPS non-numeric values\n")
                #print(refgps)
                refgps_array = np.array(refgps)
                refgps_array = refgps_array.astype(int)

                mean = np.mean(refgps_array)
                stdev.append(mean)
                self.PseudoOutput.setText(f"Average {a+1} day is : {mean}\n")

                #print ("Transfering STD data to Excel File . . .\n")
                cell = worksheet.cell(row=2,column=2+a)
                cell.value = a+1

                for i, item in enumerate(refgps, start=3):
                    cell = worksheet.cell(row=i, column=2+a)
                    cell.value = item

            stdev = np.array(stdev)
            standar_deviasi = np.std(stdev)
            hasil = standar_deviasi/10
            hasil=f"{hasil:.2f} ns"
            self.PseudoOutput.setText(str(hasil))

            akardua = np.sqrt(2)

            sum = workbook.create_sheet(title=f" Data CGGTTS ")

            cell = sum.cell (row=2, column=2)
            cell.value = "Rata-Rata Setiap Bulan"

            cell = sum.cell (row=2, column=4)
            cell.value = "Standar Deviasi"
            cell = sum.cell (row=2, column=5)
            cell.value = standar_deviasi

            for i, item in enumerate(stdev, start=3):
                cell = sum.cell(row=i, column=2+a)
                cell.value = item

            namafile = f"uPseudorange.xlsx"
            workbook.save(namafile)

        except:
            self.PseudoOutput.setText("Please Check Folder Content")


        try:
            samp = float(self.samplingnumber.text())
            sd = float(standar_deviasi/10)
            ya = sd/np.sqrt(samp)
            ups = ya*akardua
            self.uPs.setText(f"{ups:2f} ns")
        except:
            self.uPs.setText("Please input Sampling Number")



    # semuanya
    def initUI(self):
        self.setWindowTitle(' Time and Frequency Standard Remote Calibration ')

        header_layout = self.header()
        input_layout = self.input()
        run_layout = self.run()
        output_layout = self.output()
        footer_layout = self.footer()

        self.layoututama = QVBoxLayout(self)
        self.layoututama.addLayout(header_layout)
        self.layoututama.addLayout(input_layout)
        self.layoututama.addLayout(run_layout)
        self.layoututama.addLayout(output_layout)
        self.layoututama.addLayout(footer_layout)

        self.setLayout(self.layoututama)

        self.show ()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = jendelautama()

    sys.exit(app.exec_())
