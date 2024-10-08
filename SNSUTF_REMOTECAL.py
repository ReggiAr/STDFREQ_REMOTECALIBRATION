#
# The MIT License (MIT)
#
# Copyright (c) 2024 Reggi Aryunadi
# 
# This software is created for Remote Calibration services at
# the Time and Frequency Laboratory of SNSU-BSN (National Metrology Institute of Indonesia)
#
# Updated 21-08-2024 22:50 UTC(IDN)

'''
SOFTWARE TIME FREQUENCY REMOTE CALIBRATION
CGGTTS ANALYZER
UPDATE 10/09/2024
'''

# -------- Libraries

# -------- -------- System Libraries
import sys
import os

# -------- -------- GUI Libraries
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QProgressBar, QTextEdit, QDialog
from PyQt5.QtWidgets import QFileDialog, QHBoxLayout, QLineEdit, QPushButton, QComboBox, QMessageBox, QSpacerItem, QSizePolicy
from PyQt5.QtGui import QFont, QIcon, QDesktopServices
from PyQt5.QtCore import Qt, QUrl, pyqtSignal

# -------- -------- Time Libraries
from datetime import date
from astropy.time import Time
import time
from datetime import datetime, timedelta

# -------- -------- Time Libraries
from openpyxl import Workbook
import numpy as np

# -------- -------- BIPM CIRT
import requests

# -------- GUI Design

# -------- -------- Color
oren = "background-color: #369FFF; color: white"
ijau = "background-color: #5CA904; color: white"
biu = "background-color: #FFA500; color: white"
prog = "QProgressBar { border-radius :8px ; text-align: center; background-color:#369FFF ; color: white; border: 1px solid black;} QProgressBar::chunk { background-color:#5CA904 ; border-radius :8px; }"

# -------- -------- Font
font = QFont("Inter",10)

# Class
class jendelautama(QWidget):
    def __init__(self):
        super().__init__()

        self.mjdA = []
        self.aver = []

        self.utcidn = []
        self.uutdata = []

        self.initUI()
    
    # ini berisi tampilan untuk bagian header
    def header(self):

        judul = QLabel(self)
        judul.setText("  Time and Frequency Remote Calibration Software\t\t\t\t\n  CGGTTS Analyzer")
        judul.setStyleSheet(f"{oren}; font-weight: DemiBold")
        judul.setFont(QFont("Inter", 12))

        spasi = QSpacerItem(60, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)

        self.informasi = QPushButton(self)
        self.informasi.setText(" i ")
        self.informasi.setStyleSheet(oren)
        self.informasi.setFont(font)
        self.informasi.clicked.connect(self.open_pdf)

        # set tanggal hari
        hari_ini = date.today()
        tanggal_iso = hari_ini.isoformat()
        time_obj = Time(tanggal_iso, format='iso')
        tanggal_terformat = hari_ini.strftime("%d-%m-%Y")
        mjd = time_obj.mjd
        mjd_str = str(mjd)
        mjd_str = round(mjd)
        mjd_int = int(mjd_str)

        tanggalBiasa = QLabel(self)
        tanggalBiasa.setText(f"   {tanggal_terformat}   ")
        tanggalBiasa.setFont(font)
        tanggalBiasa.setStyleSheet(oren)

        tanggalMJD = QLabel(self)
        tanggalMJD.setText(f"    {mjd_int}    ")
        tanggalMJD.setFont(font)
        tanggalMJD.setStyleSheet(oren)

        satu = QHBoxLayout ()
        satu.addWidget(tanggalBiasa)
        satu.addWidget(tanggalMJD)
        satu.addItem(spasi)
        satu.addItem(spasi)
        satu.addWidget(self.informasi)

        headerLayout = QVBoxLayout ()
        headerLayout.addWidget(judul)
        headerLayout.addLayout(satu)
        headerLayout.addItem(spasi)

        return headerLayout
    
    # Untuk input input
    def input (self):

        nama = QLabel("Client Name")
        nama.setFont(font)
        tanggal = QLabel("MJD")
        tanggal.setFont(font)

        self.dirStandar = QPushButton (self)
        self.dirStandar.setText("Directory Standard")
        self.dirStandar.setFont(font)
        self.dirStandar.setStyleSheet(oren)
        self.dirStandar.clicked.connect(self.stand)

        self.dirUUT = QPushButton (self)
        self.dirUUT.setText("Directory UUT")
        self.dirUUT.setFont(font)
        self.dirUUT.setStyleSheet(oren)
        self.dirUUT.clicked.connect(self.uuts)

        self.dirOutput = QPushButton (self)
        self.dirOutput.setText("Directory Output")
        self.dirOutput.setFont(font)
        self.dirOutput.setStyleSheet(oren)
        self.dirOutput.clicked.connect(self.outputss)

        self.locStandar = QLineEdit(self)
        self.locStandar.setFont(font)
        self.locStandar.setReadOnly(True)
        self.locUUT = QLineEdit(self)
        self.locUUT.setFont(font)
        self.locUUT.setReadOnly(True)
        self.locOutput = QLineEdit(self)
        self.locOutput.setFont(font)
        self.locOutput.setReadOnly(True)

        self.tipeStandar = QComboBox(self)
        self.tipeStandar.setFont(font)
        self.tipeStandar.setStyleSheet(oren)
        self.tipeStandar.addItems(["01","2E"])

        self.tipeUUT = QComboBox(self)
        self.tipeUUT.setFont(font)
        self.tipeUUT.setStyleSheet(oren)
        self.tipeUUT.addItems(["01","2E"])

        spasi = QSpacerItem(60, 40, QSizePolicy.Expanding, QSizePolicy.Minimum)
        
        satu = QVBoxLayout()
        satu.addWidget(self.dirStandar)
        satu.addWidget(self.dirUUT)
        satu.addWidget(self.dirOutput)

        dua = QVBoxLayout()
        dua.addWidget(self.locStandar)
        dua.addWidget(self.locUUT)
        dua.addWidget(self.locOutput)

        tiga = QVBoxLayout()
        tiga.addWidget(self.tipeStandar)
        tiga.addWidget(self.tipeUUT)
        tiga.addItem(spasi)

        inputsLayout = QHBoxLayout()
        inputsLayout.addLayout(satu)
        inputsLayout.addLayout(dua)
        inputsLayout.addLayout(tiga)

        inputsLayout.setStretchFactor(satu,1)
        inputsLayout.setStretchFactor(dua,5)
        inputsLayout.setStretchFactor(tiga,1)

        return inputsLayout

    # Untuk running
    def run (self):
        satu = QHBoxLayout()
        dua = QHBoxLayout()
        tiga = QHBoxLayout()
        empat = QHBoxLayout()

        cl = QLabel()
        cl.setText("Client Name")
        cl.setFont(font)

        mjdl = QLabel()
        mjdl = QLabel()
        mjdl.setText("MJD")
        mjdl.setFont(font)

        utclabel = QLabel()
        utclabel.setText("UTC(IDN)")
        utclabel.setAlignment(Qt.AlignCenter)
        utclabel.setFont(font)
        
        uutlabel = QLabel()
        uutlabel.setAlignment(Qt.AlignCenter)
        uutlabel.setText("UUT")
        uutlabel.setFont(font)

        self.clientname = QLineEdit(self)
        self.mjdname = QLineEdit(self)
        self.clientname.setFont(font)
        self.mjdname.setFont(font)

        self.kor = QLineEdit(self)
        self.kor.setFont(font)
        self.kor.setStyleSheet(oren)
        self.kor.setReadOnly(True)

        self.utcname = QComboBox(self)
        self.utcname.setStyleSheet(oren)
        self.utcname.setFont(font)
        self.uutname = QComboBox(self)
        self.uutname.setStyleSheet(oren)
        self.uutname.setFont(font)

        self.loading = QProgressBar (self)
        self.loading.setStyleSheet(prog)
        self.loading.setFont(font)
        
        spasi = QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.analize = QPushButton(self)
        self.analize.setText("ANALYZE")
        self.analize.setFont(font)
        self.analize.setStyleSheet(oren)
        self.analize.clicked.connect(self.cirt)

        self.done = QPushButton(self)
        self.done.setText("DONE")
        self.done.setFont(font)
        self.done.setStyleSheet(ijau)
        self.done.clicked.connect(self.selesai)

        satu.addWidget(cl)
        satu.addWidget(self.clientname)
        satu.addWidget(mjdl)
        satu.addWidget(self.mjdname)
        satu.addWidget(self.kor)

        satu.setStretchFactor(cl,1)
        satu.setStretchFactor(self.clientname,2)
        satu.setStretchFactor(mjdl,1)
        satu.setStretchFactor(self.mjdname,2)
        satu.setStretchFactor(self.kor,1)

        dua.addWidget(utclabel)
        dua.addWidget(uutlabel)

        tiga.addWidget(self.utcname)
        tiga.addWidget(self.uutname)

        empat.addWidget(self.loading)
        empat.addWidget(self.analize)
        empat.addWidget(self.done)

        run_layout = QVBoxLayout()
        run_layout.addItem(spasi)
        run_layout.addLayout(satu)
        run_layout.addLayout(dua)
        run_layout.addLayout(tiga)
        run_layout.addLayout(empat)

        return run_layout

    # untuk output
    def output (self):

        corrlabel = QLabel()
        corrlabel.setText("Correction")
        corrlabel.setFont(font)

        outlabel = QLabel()
        outlabel.setText("O U T P U T")
        outlabel.setFont(font)

        #self.allan = QLineEdit(self)
        self.correction = QLineEdit(self)
        #self.allan.setFont(font)
        self.correction.setFont(font)
        #self.allan.setReadOnly(True)
        self.correction.setReadOnly(True)

        self.outputs = QTextEdit(self)
        self.outputs.setFont(font)

        self.delete = QPushButton(self)
        self.delete.setText("Delete")
        self.delete.setFont(font)
        self.delete.setStyleSheet(oren)
        self.delete.clicked.connect(self.hapus)

        self.uPsudorange = QPushButton(self)
        self.uPsudorange.setText("uPsudorange")
        self.uPsudorange.setFont(font)
        self.uPsudorange.setStyleSheet(ijau)
        self.uPsudorange.clicked.connect(self.uPseudorange)

        self.mjdcal = QPushButton(self)
        self.mjdcal.setText("MJD Calc")
        self.mjdcal.setFont(font)
        self.mjdcal.setStyleSheet(ijau)
        self.mjdcal.clicked.connect(self.mjdCalculator)

        self.klikpolar = QPushButton(self)
        self.klikpolar.setText("UTC(IDN) Data")
        self.klikpolar.setStyleSheet(biu)
        self.klikpolar.setFont(font)
        self.klikpolar.clicked.connect(self.polarx5)

        spasi = QSpacerItem(60, 50, QSizePolicy.Expanding, QSizePolicy.Minimum)

        satu = QVBoxLayout()
        dua = QVBoxLayout()
        untukoutput = QHBoxLayout()

        #satu.addWidget(alllabel)
        #satu.addWidget(self.allan)
        satu.addWidget(corrlabel)
        satu.addWidget(self.correction)
        satu.addWidget(self.delete)
        satu.addItem(spasi)
        satu.addWidget(self.mjdcal)
        satu.addWidget(self.uPsudorange)
        satu.addWidget(self.klikpolar)

        dua.addWidget(outlabel)
        dua.addWidget(self.outputs)

        untukoutput.addLayout(satu)
        untukoutput.addLayout(dua)

        untukoutput.setStretchFactor(satu,1)
        untukoutput.setStretchFactor(dua,5)

        return untukoutput

    # footer
    def footer (self):
        foot = QVBoxLayout()

        spasi = QSpacerItem(30, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        kaki = QLabel()
        kaki.setText("CGGTTS Analyzer for Cesium Atomic Clock Remote Calibration © SNSU Time and Frequency Laboratory 2024 ")

        foot.addItem(spasi)
        foot.addWidget(kaki)

        return foot

    #---------------------- uPseudorange
    #GUI
    def uPseudorange (self):
        uPseudoBox = QDialog(self)
        uPseudoBox.setWindowTitle("uPseudorange Calculator")

        judul = QLabel(uPseudoBox)
        judul.setText(" uPseudorange Calculator\t")
        judul.setFont(font)
        judul.setStyleSheet(f"{oren}; font-weight: Bold")

        tutor = QLabel(uPseudoBox)
        tutor.setText("Prepare a folder containing CGGTTS data for a full month in .txt format, with files named from 1 to 30/31")
        tutor.setFont(font)
        tutor.setStyleSheet(oren)

        cfolder = QPushButton(uPseudoBox)
        cfolder.setText("Choose Folder")
        cfolder.setFont(font)
        cfolder.setStyleSheet(oren)
        cfolder.clicked.connect(self.showDialog)

        self.foldir = QLineEdit(uPseudoBox)
        self.foldir.setFont(font)
        self.foldir.setReadOnly(True)

        jumlahlabel = QLabel(uPseudoBox)
        jumlahlabel.setText("Total Days")
        jumlahlabel.setFont(font)

        self.jumlah = QComboBox(uPseudoBox)
        self.jumlah.addItems(["28","29","30","31"])
        self.jumlah.setFont(font)
        self.jumlah.setStyleSheet(oren)

        cal = QPushButton(uPseudoBox)
        cal.setText("Calculate")
        cal.setFont(font)
        cal.setStyleSheet(oren)
        cal.clicked.connect(self.calUPseudo)

        self.PseudoOutput = QLineEdit(uPseudoBox)
        self.PseudoOutput.setFont(font)
        self.PseudoOutput.setStyleSheet(oren)
        self.PseudoOutput.setReadOnly(True)

        snumber = QLabel(uPseudoBox)
        snumber.setText("Sampling Number")
        snumber.setFont(font)

        self.samplingnumber = QLineEdit(uPseudoBox)
        self.samplingnumber.setFont(font)

        o = QLabel(uPseudoBox)
        o.setText("|")
        o.setFont(font)

        slabel = QLabel(uPseudoBox)
        slabel.setText("STDEV")
        slabel.setFont(font)

        ulabel = QLabel(uPseudoBox)
        ulabel.setText("uPseudorange")
        ulabel.setFont(font)

        self.uPs = QLineEdit(uPseudoBox)
        self.uPs.setFont(font)
        self.uPs.setStyleSheet(oren)
        self.uPs.setReadOnly(True)

        spasi = QSpacerItem(30, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        kaki = QLabel(uPseudoBox)
        kaki.setText("CGGTTS Analyzer for Cesium Atomic Clock Remote Calibration © SNSU Time and Frequency Laboratory 2024 ")
        kaki.setFont(font)

        satu = QHBoxLayout()
        satu.addWidget(cfolder)
        satu.addWidget(self.foldir)
        satu.addWidget(jumlahlabel)
        satu.addWidget(self.jumlah)
        satu.setStretchFactor(cfolder,1)
        satu.setStretchFactor(self.foldir,5)
        satu.setStretchFactor(jumlahlabel,1)
        satu.setStretchFactor(self.jumlah,1)

        dua = QHBoxLayout()
        dua.addWidget(cal)
        dua.addWidget(snumber)
        dua.addWidget(self.samplingnumber)
        dua.addWidget(o)
        dua.addWidget(slabel)
        dua.addWidget(self.PseudoOutput)
        dua.addWidget(ulabel)
        dua.addWidget(self.uPs)

        layout = QVBoxLayout(uPseudoBox)
        layout.addWidget(judul)
        layout.addItem(spasi)
        layout.addWidget(tutor)
        layout.addItem(spasi)
        layout.addLayout(satu)
        layout.addLayout(dua)
        layout.addItem(spasi)
        layout.addWidget(kaki)

        uPseudoBox.show()

    def mjdCalculator (self):

        spasi = QSpacerItem(30, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        mjdCalculator = QDialog(self)
        mjdCalculator.setWindowTitle("MJD DATE CALCULATOR")

        mjd = QLabel(self)
        mjd.setFont(font)
        mjd.setText("MJD\t")

        tanggal = QLabel(self)
        tanggal.setFont(font)
        tanggal.setText("Date\t")

        bulan = QLabel(self)
        bulan.setFont(font)
        bulan.setText("Month\t")

        tahun = QLabel(self)
        tahun.setFont(font)
        tahun.setText("Year\t")

        self.tomb1 = QPushButton(self)
        self.tomb1.setText("Calculate MJD")
        self.tomb1.setFont(font)
        self.tomb1.setStyleSheet(ijau)
        self.tomb1.clicked.connect(self.datetomjd)

        self.tanggals = QLineEdit(self)
        self.tanggals.setStyleSheet(oren)
        self.tanggals.setFont(font)
        
        self.bulans = QLineEdit(self)
        self.bulans.setStyleSheet(oren)
        self.bulans.setFont(font)

        self.tahuns = QLineEdit(self)
        self.tahuns.setStyleSheet(oren)
        self.tahuns.setFont(font)

        self.mjds = QLineEdit(self)
        self.mjds.setStyleSheet(oren)
        self.mjds.setFont(font)

        self.tomb2 = QPushButton(self)
        self.tomb2.setText("Calculate DATE")
        self.tomb2.setFont(font)
        self.tomb2.setStyleSheet(ijau)
        self.tomb2.clicked.connect(self.mjdtodate)

        satu = QVBoxLayout ()
        satu.addWidget(mjd)
        satu.addWidget(tanggal)
        satu.addWidget(bulan)
        satu.addWidget(tahun)

        dua = QVBoxLayout ()
        dua.addWidget(self.mjds)
        dua.addWidget(self.tanggals)
        dua.addWidget(self.bulans)
        dua.addWidget(self.tahuns)

        lima = QHBoxLayout()
        lima.addItem(spasi)
        lima.addItem(spasi)
        lima.addWidget(self.tomb1)
        lima.addWidget(self.tomb2)

        tiga = QHBoxLayout()
        tiga.addLayout(satu)
        tiga.addItem(spasi)
        tiga.addLayout(dua)

        empat = QVBoxLayout (mjdCalculator)
        empat.addLayout(tiga)
        empat.addLayout(lima)
        empat.addItem(spasi)

        mjdCalculator.show()

    def datetomjd (self):
        tanggal = int(self.tanggals.text())
        bulan = int(self.bulans.text())
        tahun = int(self.tahuns.text())

        date = datetime(tahun,bulan,tanggal)
        mjd_epoch = datetime(1858, 11, 17)

        delta = date - mjd_epoch
        hasil = delta.days
        hasils = str(hasil)
        self.mjds.setText(hasils)

    def mjdtodate (self):
        mjd = int(self.mjds.text())
        mjd_epoch = datetime(1858, 11, 17)

        date = mjd_epoch + timedelta(days=mjd)

        tanggal = str(date.day)
        tahun = str (date.year)
        bulan = str (date.month)

        self.tanggals.setText(tanggal)
        self.bulans.setText(bulan)
        self.tahuns.setText(tahun)

        
    #akses polar x 5 tr
    def polarx5 (self):
        url = QUrl('http://10.5.4.242/')  # Ganti dengan URL yang diinginkan
        QDesktopServices.openUrl(url)

    #----------Buka File
    def showDialog(self):
        # Buka dialog pemilihan folder
        folder = QFileDialog.getExistingDirectory(self, 'Pilih Folder')

        # Jika pengguna memilih folder, tampilkan alamat folder di QLineEdit
        if folder:
            self.foldir.setText(folder)

    # ----------- Perhitungan
    def calUPseudo(self):
        
        days = self.jumlah.currentText()
        day = int(days)
        worksheet = self.workbook.create_sheet(title=f" Data CGGTTS ")

        stdev = []
        try:
            for a in range (day):
                refgps = []
                self.PseudoOutput.setText(f"Calculate {a+1} day")
                #Read File
                with open(f'{self.foldir.text()}/{a+1}.txt', 'r') as Filestd:
                    # Baca semua baris kecuali baris pertama
                    cggtts = Filestd.readlines()

                nama_kolom = ['SAT','STTIME','REFSYS']
                header = cggtts[17].strip().split()
                std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]

                # Loop melalui setiap baris file
                for line in cggtts[19:]:  # Mulai dari baris kedelapan belas karena baris pertama adalah header
                    data = line.strip().split()
        
                    # Pastikan jumlah elemen dalam baris sesuai dengan jumlah kolom yang diharapkan
                    if len(data) >= max(std_indeks_kolom) + 1:

                        # Masukkan ke dalam list yang sesuai
                        if data[std_indeks_kolom[2]].isdigit() or data[std_indeks_kolom[2]].startswith("-") or data[std_indeks_kolom[2]].startswith("+"):
                            refgps.append(data[std_indeks_kolom[2]])
                        else:
                            print("Skipping stdRefGPS non-numeric values\n")
                #print(refgps)
                refgps_array = np.array(refgps)
                refgps_array = refgps_array.astype(int)

                mean = np.mean(refgps_array)
                stdev.append(mean)
                self.PseudoOutput.setText(f"Average {a+1} day is : {mean}\n")

                #print ("Transfering STD data to Excel File . . .\n")
                cell = worksheet.cell(row=2,column=2+a)
                cell.value = a+1

                for i, item in enumerate(refgps, start=3):
                    cell = worksheet.cell(row=i, column=2+a)
                    cell.value = item

            stdev = np.array(stdev)
            standar_deviasi = np.std(stdev)
            hasil = standar_deviasi/10
            hasil=f"{hasil:.2f} ns"
            self.PseudoOutput.setText(str(hasil))

            akardua = np.sqrt(2)

            sum = self.workbook.create_sheet(title=f" Data CGGTTS ")

            cell = sum.cell (row=2, column=2)
            cell.value = "Rata-Rata Setiap Bulan"

            cell = sum.cell (row=2, column=4)
            cell.value = "Standar Deviasi"
            cell = sum.cell (row=2, column=5)
            cell.value = standar_deviasi

            for i, item in enumerate(stdev, start=3):
                cell = sum.cell(row=i, column=2+a)
                cell.value = item

            namafile = f"uPseudorange.xlsx"
            self.workbook.save(namafile)

        except:
            self.PseudoOutput.setText("Please Check Folder Content")

        try:
            samp = float(self.samplingnumber.text())
            sd = float(standar_deviasi/10)
            ya = sd/np.sqrt(samp)
            ups = ya*akardua
            self.uPs.setText(f"{ups:2f} ns")
        except:
            self.uPs.setText("Please input Sampling Number")

    # -------------- Ini Mulai Fungsi Fungsi
    # klik standard
    def stand (self):
        # Buka dialog pemilihan folder
        folder = QFileDialog.getExistingDirectory(self, 'Pilih Folder')

        # Jika pengguna memilih folder, tampilkan alamat folder di QLineEdit
        if folder:
            self.locStandar.setText(folder)
        
        self.utcname.clear()
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                self.utcname.addItem(filename)

    #buka file pdf
    def open_pdf(self):
        file_path ='I.MF.2.04 Remote Calibration.pdf'
        if os.path.exists(file_path):
            os.startfile(file_path)

    # klik uut
    def uuts (self):
        # Buka dialog pemilihan folder
        folder = QFileDialog.getExistingDirectory(self, 'Pilih Folder')

        # Jika pengguna memilih folder, tampilkan alamat folder di QLineEdit
        if folder:
            self.locUUT.setText(folder)

        self.uutname.clear()
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                self.uutname.addItem(filename)
    
    # klik output
    def outputss (self):
        # Buka dialog pemilihan folder
        folder = QFileDialog.getExistingDirectory(self, 'Pilih Folder')

        # Jika pengguna memilih folder, tampilkan alamat folder di QLineEdit
        if folder:
            self.locOutput.setText(folder)
        
# -------- DISINI PERHITUNGANNYA DIMULAI -----------------------------------------------

    # koreksi connect ke BIPM
    def find_numbers(self,n):
        try:
            if str(n).endswith('4') or str(n).endswith('9'):
                r = requests.get(f"https://webtai.bipm.org/api/v0.2-beta/get-data.html?scale=utc&lab=IDN&outfile=txt&&mjd1={n}&mjd2={n}")
                data = r.text.split('\n')
                for line in data:
                    if line and "UTC-UTC(IDN)(ns)" not in line:
                        parts = line.split()  # Split the line into parts
                        if len(parts) > 1:  # Ensure there are at least two parts
                            try:
                                value = float(parts[1])  # The value is at index 1
                                hasil = str(f"{value:.2f}")
                                self.kor.setText(hasil)
                                #print(f"UTC-UTC(IDN) at {n} is : {value}")
                            except ValueError:
                                print("Error converting to float")

            else:
                before = n - 1
                after = n + 1

                while not (str(before).endswith('4') or str(before).endswith('9')):
                    before -= 1

                while not (str(after).endswith('4') or str(after).endswith('9')):
                    after += 1

                r = requests.get(f"https://webtai.bipm.org/api/v0.2-beta/get-data.html?scale=utc&lab=IDN&outfile=txt&&mjd1={before}&mjd2={after}")
                data = r.text.split('\n')
                values = []
                for item in data:
                    parts = item.split()
                    if len(parts) == 2:
                        try:
                            value = float(parts[1])
                            values.append(value)
                        except ValueError:
                            pass
                            #print("Error converting to float")

                if len(values) >= 2:
                    kons = (values[1]-values[0])/5
                    #print(kons)
                    nilai = values[0]+kons*(n-before)
                    hasil = str(f"{nilai:.2f}")
                    self.kor.setText(hasil)
                    #print(f"UTC - UTC(IDN) at {n} is {nilai}")
                else:
                    print("Not enough data found.")
        except :
            self.kor.setText("Not Found")

    # BACA FILE TXT
    def readprn (self, stdFile, Format):
        Prn = []

        with open(f'{stdFile}','r') as Filestd:
            stdLines = Filestd.readlines()

        if Format == "2E":
            try:
                nama_kolom = ['SAT','STTIME','REFSYS']
                header = stdLines[17].strip().split()
                std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]
            except:
                print("17 not found")
        else:
            nama_kolom = ['PRN','STTIME','REFGPS']
            header = stdLines[17].strip().split()
            std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]
        
        for line in stdLines[19:]:
            data = line.strip().split()
            if len(data) >= max(std_indeks_kolom) + 1:
                if Format == "2E":
                    Prn.append(data[std_indeks_kolom[0]])
                else:
                    if data[std_indeks_kolom[0]].isdigit():
                        Prn.append(data[std_indeks_kolom[0]])
                    else:
                        print("Skipping stdprn non-numerical values")
        

        # HAPUSIN HURUF DEPAN
        if Format == "2E":
            bersih =[]
            for item in Prn:
                if item[0].isdigit():
                    bersih.append(item)
                else:
                    bersih.append(item[1:])
            #print(bersih)
            Prn = []

            for i, item in enumerate(bersih):
                Prn.append(bersih[i])
        
        return Prn

    def readsttime (self, stdFile, Format):
        Sttime = []

        with open(f'{stdFile}','r') as Filestd:
            stdLines = Filestd.readlines()

        if Format == "2E":
            try:
                nama_kolom = ['SAT','STTIME','REFSYS']
                header = stdLines[17].strip().split()
                std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]
            except:
                print("17 not found")

        else:
            nama_kolom = ['PRN','STTIME','REFGPS']
            header = stdLines[17].strip().split()
            std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]

        for line in stdLines[19:]:
            data = line.strip().split()
            if len(data) >= max(std_indeks_kolom) + 1:
                if data[std_indeks_kolom[1]].isdigit():
                    Sttime.append(data[std_indeks_kolom[1]])
                else:
                    print("Skipping stdttime non-numeric values")
        return Sttime

    def readRefGPS (self, stdFile, Format):  
        RefGPS = []

        with open(f'{stdFile}','r') as Filestd:
            stdLines = Filestd.readlines()

        if Format == "2E":
            try:
                nama_kolom = ['SAT','STTIME','REFSYS']
                header = stdLines[17].strip().split()
                std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]
            except:
                print("17 not found")

        else:
            nama_kolom = ['PRN','STTIME','REFGPS']
            header = stdLines[17].strip().split()
            std_indeks_kolom = [header.index(kolom) for kolom in nama_kolom]

        for line in stdLines[19:]:
            data = line.strip().split()

            if len(data) >= max(std_indeks_kolom) + 1:           
                
                if data[std_indeks_kolom[2]].isdigit() or data[std_indeks_kolom[2]].startswith("-") or data[std_indeks_kolom[2]].startswith("+"):
                    RefGPS.append(data[std_indeks_kolom[2]])
                else:
                    print("Skipping stdRefGPS non-numeric values\n")

        return RefGPS

    def excel (self, posisi, data, mjd):

        for i, item in enumerate (data,start=4):
            cell = mjd.cell(row=i, column=posisi)
            cell.value=item

    def cek (self, ini):
        for a in range(len(ini) - 1):
            if ini[a] == ini[a-1]:
                
                ini[a] += 1
            else:
                pass
        return ini
    
    def allan_variance(self,data, tau):
        """
        Menghitung Allan Variance dari data list.

        Parameters:
        data (list atau array): List atau array yang berisi data numerik.
        tau (int): Interval waktu (atau ukuran grup) yang digunakan untuk menghitung Allan Variance. Default-nya adalah 1.

        Returns:
        float: Allan Variance dari data.
        """
        try:
            ini = np.array(data)
            n = len(ini)
            if n < 2 * tau:
                raise ValueError("Data tidak cukup untuk menghitung Allan Variance dengan nilai tau yang diberikan.")

            # Membuat list rata-rata per grup
            group_means = [np.mean(data[i:i+tau]) for i in range(0, n - tau, tau)]
        
            # Menghitung Allan Variance
            squared_diffs = [(group_means[i+1] - group_means[i])**2 for i in range(len(group_means) - 1)]
        
            allan_var = 0.5 * np.mean(squared_diffs)
        except:
            allan_var="0"
    
        return allan_var
    
    def show_warning(self, pesan):
        # Membuat jendela peringatan
        warning = QMessageBox()
        warning.setIcon(QMessageBox.Warning)
        warning.setWindowTitle('Peringatan')
        warning.setText(pesan)
        
        warning.exec_()

    # Dimulai disini
    # 1. Request data circular-T dulu
    def cirt (self):
        
        a = self.dirStandar.text()
        b = self.dirUUT.text()
        c = self.dirOutput.text()
        d = self.clientname.text()
        e = self.mjdname.text()
        f = self.utcname.currentText()
        g = self.uutname.currentText()

        if a == "":
            self.show_warning("Anda Belum Memilih Directory Folder Standard")
        elif b =="":
            self.show_warning("Anda Belum Memilih Directory Folder UUT")
        elif c =="":
            self.show_warning("Anda Belum Memilih Directory Folder Output")
        elif d =="":
            self.show_warning("Anda Belum Mengisi Nama Client")
        elif e =="":
            self.show_warning("Anda Belum Mengisi Tanggal MJD")
        elif not f.endswith('.txt'):
            self.show_warning("File Standard yang dipilih bukan berformat .txt")
        elif not g.endswith('.txt'):
            self.show_warning("File UUT yang dipilih bukan berformat .txt")
        else:
            try:
                time.sleep(1)
                self.loading.setValue(10)
                try:
                    self.find_numbers(int(self.mjdname.text()))
                    time.sleep(1)
                    self.loading.setValue(20) 
                    self.read()

                except:
                    self.show_warning("Tanggal MJD Salah")

            except:
                self.show_warning("Cek File Excel")

    # 2. Read data txt
    def read (self):
        try:
            stdfile = f"{self.locStandar.text()}/{self.utcname.currentText()}"
            stdformats = self.tipeStandar.currentText()

            uutfile = f"{self.locUUT.text()}/{self.uutname.currentText()}"
            uutformats = self.tipeUUT.currentText()
            
            time.sleep(1)
            self.loading.setValue(30)

        except:
            self.show_warning("Error saat membaca file txt")

        self.stdPrn = self.readprn(stdfile,stdformats)
        self.stdSttime= self.readsttime(stdfile,stdformats)
        stdRefGPS = self.readRefGPS(stdfile,stdformats)
        stdrefgpss = [float(x) for x in stdRefGPS]
        self.stdRefGPS = [num / 10 for num in stdrefgpss]

        self.uutPrn = self.readprn(uutfile, uutformats)
        self.uutSttime = self.readsttime(uutfile, uutformats)
        uutRefGPS = self.readRefGPS(uutfile, uutformats)
        uutRefGPSS = [float(x) for x in uutRefGPS]
        self.uutRefGPS = [num/10 for num in uutRefGPSS]

        time.sleep(5)
        self.loading.setValue(40)

        self.koreksi()

    # 3. Koreksi dengan data circular-T
    def koreksi(self):
        korek = (float(self.kor.text()))

        self.stdcorrefgps = list(map(lambda x: x - abs(korek), self.stdRefGPS))

        time.sleep(1)
        self.loading.setValue(50)

        self.ref()

    # 4. Bikin Ref
    def ref(self):
        a = [float(x) for x in self.stdPrn]
        b = [float(x) for x in self.stdSttime]

        self.stdRefVal = list(map(lambda x , y: x * y, a,b))

        c = [float(x) for x in self.uutPrn]
        d = [float(x) for x in self.uutSttime]

        self.uutRefVal = list(map(lambda x , y: x * y, c,d))

        time.sleep(1)
        self.loading.setValue(60)

        self.sorting()

    # 5. Sorting
    def sorting (self):
        std = list(zip(self.stdRefVal, self.stdcorrefgps))
        uut = list(zip(self.uutRefVal, self.uutRefGPS))

        sort_std = sorted(std, key=lambda x: x[0])
        sort_uut = sorted(uut, key=lambda x: x[0])

        sort_std_refval = [data[0] for data in sort_std]
        self.sort_std_refgps = [data[1] for data in sort_std]

        sort_uut_refval = [data[0] for data in sort_uut]
        self.sort_uut_refgps = [data[1] for data in sort_uut]

        self.sort_std_refval = self.cek(sort_std_refval)
        self.sort_uut_refval = self.cek(sort_uut_refval)

        time.sleep(1)
        self.loading.setValue(70)

        self.matched()

    # 6. Match
    def matched (self):
        self.cstdv = []
        self.cstdg = []
        self.cuutv = []
        self.cuutg = []
        for a in self.sort_std_refval:
            for b in self.sort_uut_refval:
                res = a/b
                if res == 1:
                        indexa = self.sort_std_refval.index(a)
                        indexb = self.sort_uut_refval.index(b)
                        self.cstdv.append(self.sort_std_refval[indexa])
                        self.cstdg.append(self.sort_std_refgps[indexa])
                        self.cuutv.append(self.sort_uut_refval[indexb])
                        self.cuutg.append(self.sort_uut_refgps[indexb])
                        break
                else:
                    pass

        time.sleep(1)
        self.loading.setValue(80)

        self.selisih()

    # 7. hitung selisih
    def selisih (self):
        self.beda = []
        for i in range (len(self.cstdg)):
            try:
                self.beda.append(int(self.cstdg[i])-int(self.cuutg[i]))
            except:
                print("")
        
        time.sleep(1)
        self.loading.setValue(90)

        self.conclusion()

    # 8. kesimpulan
    def conclusion (self):
        selisih = np.array(self.beda)
        sel = (float(np.mean(selisih)))
        beda = f"{sel:.2f}"
        self.outputs.append(f"Rata-rata selisih pada tanggal {self.mjdname.text()} adalah {beda}")

        utc = np.array(self.cstdg)
        utcidn = float(np.mean(utc))
        self.utcidn.append(utcidn)

        uut = np.array(self.cuutg)
        uuts = uut.astype(float)
        uutg = float(np.mean(uuts))
        self.uutdata.append(uutg)

        self.aver.append(sel)
        self.mjdA.append(int(self.mjdname.text()))

        a = np.array(self.aver)

        rerata = np.mean(a)
        rata = f"{rerata:.2f}"
        self.correction.setText(rata)

        # - - - - - - - print excel

        
        mjd = self.workbook.create_sheet(title=f"{self.mjdname.text()}")

        header1 = ["STD RAW DATA","","","","","UUT RAW DATA","","","","REF VAL","","","SORTED DATA STD","","SORTED DATA UUT","","","MATCH DATA STD","","MATCH DATA UUT","","","STD-UUT"]
        header2 = ["PRN","STTIME","REFGPS","REFGPS COR","","PRN","STTIME","REFGPS","","STD","UUT","","REFVAL","REFGPS","REFVAL","REFGPS","","REFVAL","REFGPS","REFVAL","REFGPS"]

        for i, item in enumerate (header1,start=2):
            cell = mjd.cell(row=2, column=i)
            cell.value=item

        for i, item in enumerate (header2,start=2):
            cell = mjd.cell(row=3, column=i)
            cell.value=item

        self.excel(2,self.stdPrn,mjd)
        self.excel(3,self.stdSttime,mjd)
        self.excel(4,self.stdRefGPS,mjd)
        self.excel(5, self.stdcorrefgps,mjd)

        self.excel(7,self.uutPrn,mjd)
        self.excel(8,self.uutSttime,mjd)
        self.excel(9,self.uutRefGPS,mjd)

        self.excel (11, self.stdRefVal,mjd)
        self.excel (12, self.uutRefVal,mjd)

        self.excel(14,self.sort_std_refval,mjd)
        self.excel(15,self.sort_std_refgps,mjd)

        self.excel(16,self.sort_uut_refval,mjd)
        self.excel(17,self.sort_uut_refgps,mjd)

        self.excel(19, self.cstdv,mjd)
        self.excel(20, self.cstdg,mjd)
        self.excel(21, self.cuutv,mjd)
        self.excel(22, self.cuutg,mjd)

        self.excel(24,self.beda,mjd)

        excelFile = f"{self.locOutput.text()}/{self.clientname.text()}.xlsx"
        self.workbook.save(excelFile)

        time.sleep(1)
        tambah = int(self.mjdname.text())+1
        self.mjdname.setText(str(tambah))
        self.loading.setValue(100)

    # 9. Done
    def selesai (self):
        summary = self.workbook.create_sheet(title=f" Summary ")

        for i, item in enumerate (self.mjdA,start=4):
            cell = summary.cell(row=i, column=2)
            cell.value=item
        for i, item in enumerate (self.utcidn,start=4):
            cell = summary.cell(row=i, column=4)
            cell.value=item
        for i, item in enumerate (self.uutdata,start=4):
            cell = summary.cell(row=i, column=5)
            cell.value=item
        for i, item in enumerate (self.aver,start=4):
            cell = summary.cell(row=i, column=6)
            cell.value=item

        cell = summary.cell(row=3,column=2)
        cell.value = "MJD"
        cell = summary.cell(row=3,column=4)
        cell.value = "UTC(IDN)"
        cell = summary.cell(row=3,column=5)
        cell.value = "UUT"
        cell = summary.cell(row=3,column=6)
        cell.value = "UTC(IDN) - UUT"

        excelFile = f"{self.locOutput.text()}/{self.clientname.text()}.xlsx"
        self.workbook.save(excelFile)
        self.show_warning(f"Worksheet anda telah tersimpan di {excelFile}")
    
    # 10. Hapus
    def hapus (self):
        try:
            self.aver.pop()
            self.mjdA.pop()
        except:
            self.show_warning("Jangan Kebanyakan Hapusnya")

        a = np.array(self.aver)
        beda = np.mean(a)
        self.correction.setText(f"{beda:.2f}")

        cursor = self.outputs.textCursor()
        cursor.movePosition(cursor.End)  # Pindah ke akhir teks
        cursor.movePosition(cursor.StartOfBlock, cursor.KeepAnchor)  # Pilih baris terakhir
        cursor.removeSelectedText()  # Hapus teks yang dipilih
        cursor.deleteChar()  # Hapus karakter newline jika ada
        self.outputs.setTextCursor(cursor)

    # semuanya
    def initUI(self):
        self.setWindowTitle(' Time and Frequency Standard Remote Calibration ')

        ikon = QIcon(r'icon.png')
        self.setWindowIcon(ikon)

        header_layout = self.header()
        input_layout = self.input()
        run_layout = self.run()
        output_layout = self.output()
        footer_layout = self.footer()

        self.layoututama = QVBoxLayout(self)
        self.layoututama.addLayout(header_layout)
        self.layoututama.addLayout(input_layout)
        self.layoututama.addLayout(run_layout)
        self.layoututama.addLayout(output_layout)
        self.layoututama.addLayout(footer_layout)

        self.setLayout(self.layoututama)
    
        self.workbook = Workbook()

        self.show ()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = jendelautama()

    sys.exit(app.exec_())
